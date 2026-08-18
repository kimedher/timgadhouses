#!/usr/bin/env python3
"""Export the Timgad houses database (v16 xlsx, HOUSES sheet) to _data/houses.json.

v16 sheet layout: two title rows + section-band row, headers on row 4, one
data row per attested house from row 5. Raw columns are copied verbatim.
Derived fields (marked _derived in name comments): name_en / name_fr, zone,
type_class, area_m2.

v16 column changes handled here:
- "Alternate Names" is now "Alternate Names & Numbers".
- The French name moved to its own "French Name" column; when it is filled it
  wins over any parenthetical embedded in "House Name" (a trailing
  "(Maison ...)" duplicate is stripped from name_en; other trailing
  parentheticals are English glosses and stay). When "French Name" is empty
  the v13 rule applies: the final balanced parenthetical of "House Name"
  splits off as name_fr. A "House Name" cell holding an editorial review flag
  (starts Drop/Flag/Reconsider/Qualify) is not published; the English name is
  reconstructed from the French positional name instead.
- "Quadrant" cells may carry reconciliation prose; the leading quadrant token
  is extracted, falling back to the quadrant segment of the Grid ID.
- Area split into "Footprint L x W (m)" / "Footprint m2" / "Approx. Area";
  area_raw prefers "Approx. Area" (the published m2 figure) and falls back to
  the L x W cell, mirroring the single v13 "Approx. Area (l x w)" column.
- "Verification Status" gained UNVERIFIED, normalized to "review" alongside
  NEEDS REVIEW (the site badge vocabulary stays verified / partial / review).

The "generated" date is the database date parsed from the xlsx filename
(e.g. ..._v16_2026-08-09.xlsx), falling back to today's date.
Usage: python3 export_houses.py <xlsx path> <output json path>
"""
import csv, datetime, json, os, re, sys
import openpyxl

XLSX, OUT = sys.argv[1], sys.argv[2]

# Public-notes overlay: prose rewrites of the working notes, keyed by grid_id.
# When a grid_id appears here, its note replaces the xlsx note at export time,
# so the public site carries readable prose while the database keeps its
# working shorthand.
OVERLAY_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "public_notes.csv")
PUBLIC_NOTES = {}
if os.path.exists(OVERLAY_CSV):
    with open(OVERLAY_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            PUBLIC_NOTES[row["grid_id"].strip()] = row["note"].strip()

# Analysis-progress overlay: grid_ids with a status of "drawn" or "case_study";
# every house absent from the csv exports as "pending".
ANALYSIS_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "analysis_status.csv")
ANALYSIS_STATUS = {}
if os.path.exists(ANALYSIS_CSV):
    with open(ANALYSIS_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ANALYSIS_STATUS[row["grid_id"].strip()] = row["status"].strip()

# The 12-house dissertation sample publishes full records (notes, references).
# All other houses publish index metadata only until the analysis is defended.
SAMPLE_IDS = {
    "TIMG.E-SW.I1", "TIMG.E-SW.I2", "TIMG.SW.I5", "TIMG.E-SW.I5",
    "TIMG.NW.I6", "TIMG.E-NW.I20", "TIMG.SE.I2", "TIMG.NW.I17",
    "TIMG.SW.I8", "TIMG.E-NW.I18", "TIMG.E-NW.I19", "TIMG.NW.I25",
}

HEADER_ROW = 4  # v16: rows 1-2 titles, row 3 section bands, row 4 headers

def clean(v):
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v)).strip()

def strip_editorial_tags(s):
    # Public output stays clean: drop "[moved from ...]" editorial markers,
    # database-version tokens like ", v16", and whole working segments that are
    # addressed to the editor rather than the reader ("Correction proposed (...)",
    # "Name review (...)"). Notes are stored as " | "-separated segments, so an
    # editorial segment is dropped whole. The xlsx keeps everything; this is
    # export-side only.
    s = re.sub(r"\s*\[moved from[^\]]*\]\s*", " ", s, flags=re.IGNORECASE)
    s = re.sub(r",\s*v1\d+\]", "]", s)
    segments = [seg.strip() for seg in s.split("|")]
    segments = [seg for seg in segments if seg and not re.match(
        r"^(correction proposed|name review|correction applied|proposed correction)\b",
        seg, flags=re.IGNORECASE)]
    s = " | ".join(segments)
    return re.sub(r"\s+", " ", s).strip()

def clean_citation(s):
    # Drop internal workflow tokens from public citation fields; the xlsx
    # keeps them. Applied to first_published_by only.
    s = re.sub(
        r"\s*\((?:printed; verify in text|printed; needs scan receipt|"
        r"probable; needs scan receipt|exact page needs Kim)\)", "", s)
    s = s.replace("BIAA-Af-Alg-24, ", "").replace("(BIAA-Af-Alg-24)", "")
    s = re.sub(r"(:\s*)~(\d)", r"\1\2", s)  # "Ballu 1911: ~78-79" -> ": 78-79"
    return re.sub(r"\s+", " ", s).strip()

def split_name(name):
    # Split off the final balanced parenthetical as the French name; handles
    # nested parentheses like "House 83 (formerly 91) (Maison 83 (ex-91))".
    if name.endswith(")"):
        depth = 0
        for i in range(len(name) - 1, -1, -1):
            if name[i] == ")":
                depth += 1
            elif name[i] == "(":
                depth -= 1
                if depth == 0:
                    return name[:i].strip(), name[i + 1:-1].strip()
    return name, ""

# v16 stores editorial review flags in a few "House Name" cells; they start
# with an imperative and always sit beside a filled "French Name".
EDITORIAL_NAME = re.compile(r"(?:Drop|Flag|Reconsider|Qualify)\b")
FR_POSITIONAL = re.compile(
    r"(nord|sud)-(ouest|est) quartier, (\d+)(?:e|ère|re) rangée(?:, position (\d+))?")

def english_from_french(fr):
    # Rebuild the neutral English positional name ("House NW Quarter Row 2
    # Position 1") from the French positional name when the English cell
    # holds an editorial flag instead of a name.
    m = FR_POSITIONAL.search(fr)
    if not m: return ""
    quad = {"nord": "N", "sud": "S"}[m.group(1)] + {"ouest": "W", "est": "E"}[m.group(2)]
    s = "House %s Quarter Row %s" % (quad, m.group(3))
    if m.group(4): s += " Position %s" % m.group(4)
    return s

def names(raw, fr_col):
    # -> (name, name_en, name_fr)
    if fr_col:
        fr = fr_col
        if EDITORIAL_NAME.match(raw):
            en = english_from_french(fr_col) or fr_col
        else:
            base, emb = split_name(raw)
            # A trailing "(Maison ...)" duplicates the French Name column;
            # strip it. Other trailing parentheticals are English glosses
            # ("(Philadelphi)", "(Jardinières)") and stay in name_en.
            en = base if re.match(r"(?i)maison\b", emb) else raw
    else:
        en, fr = split_name(raw)
        # Only treat the peeled parenthetical as French when it reads like a
        # French house name; otherwise it is an English gloss or numbering
        # note ("(Insula 24)", "(Lohmann numbering)") and stays in name_en.
        if fr and not re.match(r"(?i)(maison|habitat|grande|petite|villa|demeure)\b", fr):
            en, fr = raw, ""
    name = "%s (%s)" % (en, fr) if fr else en
    return name, en, fr

QUAD_TOKEN = re.compile(r"(E-)?(NW|NE|SW|SE)\b")

def quadrant(qraw, gid):
    # The Grid ID is authoritative for the published quadrant/zone so the
    # catalog card never contradicts the ID the legend teaches readers to
    # decode. v16 Quadrant cells carry corrections pending a grid-ID
    # migration (see CHANGELOG); until the IDs are re-keyed, the cell token
    # is used only where the Grid ID itself is unlocated (TIMG.UNK.*).
    m = re.match(r"TIMG\.((?:E-)?(?:NW|NE|SW|SE))\.", gid)
    if m: return m.group(1)
    m = QUAD_TOKEN.match(qraw)
    if m: return (m.group(1) or "") + m.group(2)
    return "UNK"

def zone(q):
    if q in ("NW","NE","SW","SE"): return "Intramural " + q
    if q.startswith("E-"): return "Extramural " + q[2:]
    return "Unlocated"

def type_class(bt):
    t = bt.lower()
    if "elite" in t or "flamen" in t: return "Elite house"
    if "collective" in t or "warehouse" in t or "mixed-use" in t: return "Mixed-use / other"
    if "block" in t or "insula" in t: return "Residential block"
    if "isolated" in t: return "Freestanding house"
    if "house" in t or "residential" in t: return "House"
    return "Uncertain"

VSTATUS = {
    "VERIFIED": "verified",
    "PARTIALLY VERIFIED": "partial",
    "NEEDS REVIEW": "review",
    "UNVERIFIED": "review",  # site badge vocabulary has no fourth value
}

def verification_status(v):
    key = clean(v).upper()
    if key not in VSTATUS:
        raise ValueError(f"Unexpected Verification Status: {v!r}")
    return VSTATUS[key]

def area_m2(raw, footprint_m2):
    m = re.search(r"~?\s*([\d][\d\s.,]*)\s*m²", raw) if raw else None
    if m:
        n = m.group(1).replace(" ", "").replace(",", "")
        try: return round(float(n))
        except ValueError: pass
    if isinstance(footprint_m2, (int, float)):
        return round(footprint_m2)
    return None

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["HOUSES"]
hdr = [c.value for c in ws[HEADER_ROW]]
houses = []
for r in ws.iter_rows(min_row=HEADER_ROW + 1):
    row = dict(zip(hdr, [c.value for c in r]))
    if not any(v not in (None, "") for v in row.values()): continue
    gid = clean(row["Grid ID"])
    name, en, fr = names(clean(row["House Name"]), clean(row["French Name"]))
    q = quadrant(clean(row["Quadrant"]), gid)
    ar = clean(row["Approx. Area"]) or clean(row["Footprint L x W (m)"])
    houses.append({
        "grid_id": gid,
        "name": name,
        "name_en": en,
        "name_fr": fr,
        "alternate_names": clean(row["Alternate Names & Numbers"]),
        "quadrant": q,
        "zone": zone(q),
        "building_type": clean(row["Building Type"]),
        "type_class": type_class(clean(row["Building Type"])),
        "confidence": clean(row["Confidence"]).lower(),
        "mapping_confidence": clean(row["Mapping Confidence"]).lower(),
        "verification_status": verification_status(row["Verification Status"]),
        "analysis_status": ANALYSIS_STATUS.get(gid, "pending"),
        "area_raw": ar,
        "area_m2": area_m2(ar, row["Footprint m2"]),
        "first_published_by": clean_citation(clean(row["First Published By"])),
        "key_references": clean(row["Key References"]),
        "possible_duplicate_of": clean(row["Possible Duplicate Of"]),
        "notes": strip_editorial_tags(clean(row["Notes"])),
    })
    h = houses[-1]
    h["sample"] = h["grid_id"] in SAMPLE_IDS
    if not h["sample"]:
        h["notes"] = ""
        h["key_references"] = ""
    if h["grid_id"] in PUBLIC_NOTES:
        h["notes"] = PUBLIC_NOTES[h["grid_id"]]

houses.sort(key=lambda h: h["grid_id"])
m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(XLSX))
generated = m.group(1) if m else datetime.date.today().isoformat()
out = {"generated": generated, "count": len(houses), "houses": houses}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=1)
print("wrote", OUT, "houses:", len(houses))
import collections
print("type_class:", dict(collections.Counter(h["type_class"] for h in houses)))
print("zones:", dict(collections.Counter(h["zone"] for h in houses)))
print("areas parsed:", sum(1 for h in houses if h["area_m2"]))
print("full records (sample):", sum(1 for h in houses if h["sample"]))
